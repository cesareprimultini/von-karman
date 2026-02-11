import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from statistics import mean
import copy 
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# import pyarrow
import math
from windrose import WindroseAxes

wave_met_filepath = r"vector_decomposition\workflows\beam_workflow\P2_wave.csv"
current_met_filepath = r"vector_decomposition\workflows\beam_workflow\P2_current.csv"
# current_3d_data_filepath = r"P:\2198-CIP Feng Miao I (FM1) OWF\01 client data\Metocean Data\ECC Hindcast Data\DataExtraction\m3hd_fm1_cablecorridor_2014-2023_surf.txt"

ABSTAB_temp_sheet = r"P:\2198-CIP Feng Miao I (FM1) OWF\03 calcs\CTR 009 FLS flow amp\Stress and Fatigue Model\ABSTAB Template.xlsx"
export_path = r"vector_decomposition\workflows\beam_workflow\compiled_metocean_hindcast_2198_DRAFTNOTFINAL_CP.csv"

wave_col_names = ['DateTime', 'Hm0', 'Tp', 'Tp02', 'PWD', 'MWD']
current_col_names = ['DateTime', 'Surface Elevation', 'Speed U', 'Speed V']

wave_met_data = pd.read_csv(wave_met_filepath, skiprows = 1, sep = ',' , names = wave_col_names, header = None, encoding = 'ANSI')
current_met_data = pd.read_csv(current_met_filepath, skiprows = 1, sep = ',', names = current_col_names, header = None, encoding = 'ANSI')
# current_3d_met_data_total = pd.read_csv(current_3d_data_filepath, sep ="\t" , skiprows = [0, 2])

# current_3d_P2 = pd.DataFrame()
# current_3d_P2['Time'] = current_3d_met_data_total['Time']
# current_3d_P2['Uc'] = current_3d_met_data_total['P2: Current speed']
# current_3d_P2['heading (rads)'] = current_3d_met_data_total['P2: Current direction (Horizontal)']
# current_3d_P2['heading (deg)']  = np.degrees(current_3d_P2['heading (rads)'])

# #have read in the data
# # the data needs to be altered so that the 30 min timestamps are kept, this means that the wave data needs to be duplicated and added to match the current data for the respective hr
wave_met_data = wave_met_data.loc[wave_met_data.index.repeat(2)].reset_index(drop = True)

## old code for removing 30 min sections
# str_to_rem = ':30'

# #creating a col of true or false to see which rows to rem
# mask = current_met_data['DateTime'].str.endswith(str_to_rem)

# current_met_data = current_met_data[~mask]
# current_met_data = current_met_data.reset_index(drop = True)

# lets compare the speed U and Speed V 
plt.figure() 
plt.plot(current_met_data.index, current_met_data['Speed V'], color = 'b', label = 'speed v')
plt.plot(current_met_data.index, current_met_data['Speed U'], color = 'red', label = 'speed u')
plt.xlabel('index')
plt.ylabel('velocity (m/s)')
plt.title('Comparison of Speed U and Speed V')

# def angles_to_360(angle): 
#     if angle < 0: 
#         angle += 360
#     return angle

# ## lets test the heading calc
# for i in range(len(current_met_data)):
#     print(f'i = {i}')
#     x = current_met_data.loc[i, 'Speed U']
#     y = current_met_data.loc[i,'Speed V']
#     # print(f"x = {x}")
#     # print(f"y = {y}")
#     current_met_data.loc[i, 'Uc'] = math.sqrt((x**2) + (y**2))
#     current_met_data.loc[i, 'heading (v/u)'] = angles_to_360(math.degrees(math.atan2(y,x))) #calculates atan(y/x)
#     current_met_data.loc[i, 'heading (u/v)'] = angles_to_360(math.degrees(math.atan2(x,y))) #calculates atan(x/y)
#     # current_met_data.loc[i, 'heading'] = math.degrees(math.atan2((current_met_data.loc[i,'Speed V'])/(current_met_data.loc[i, 'Speed U'])))
#     print(f"heading for i = {i}: {current_met_data.loc[i], 'heading'}\n")
    
current_met_data['Uc'] = np.sqrt((current_met_data['Speed U'])**2 + (current_met_data['Speed V'])**2)
# current_met_data['heading (v/u)'] = np.degrees(np.arctan2(current_met_data['Speed V'], current_met_data['Speed U']))
current_met_data['heading (u/v)'] = np.degrees(np.arctan2(current_met_data['Speed U'], current_met_data['Speed V']))
#arctan gives angles from east and anticlockwise, so flipping the opposite and adjasent and %360 converts to clockwise and from north
# current_met_data['heading (v/u)'] = (current_met_data['heading (v/u)'] + 360) % 360 
current_met_data['heading (u/v)'] = (current_met_data['heading (u/v)'] + 360) % 360


ax = WindroseAxes.from_ax()
bins = [0,0.2,0.4,0.6,0.8,1.1]
# ax.bar(current_met_data['heading (v/u)'], current_met_data['Uc'], bins = bins, normed = True, opening = 1, edgecolor = 'black', nsector = 12)
ax.bar(current_met_data['heading (u/v)'], current_met_data['Uc'], bins = bins, normed = True, opening = 1, edgecolor = 'black', nsector = 12)
custom_ticks = [10, 20, 30, 40, 50]
ax.set_yticks(custom_ticks)
ax.set_yticklabels([f"{t}%" for t in custom_ticks])
ax.set_ylim(top=max(custom_ticks))
ax.set_title('Current Rose arctan2(speedU/speedV) (to shift to from north)')
ax.set_legend()

# ax = WindroseAxes.from_ax()
# bins = [0,0.2,0.4,0.6,0.8,1.1]
# ax.bar(current_3d_P2['heading (deg)'], current_3d_P2['Uc'], bins = bins,normed = True, opening = 1, edgecolor = 'black', nsector = 12)
# custom_ticks = [10, 20, 30, 40, 50]
# ax.set_yticks(custom_ticks)
# ax.set_yticklabels([f"{t}%" for t in custom_ticks])
# ax.set_ylim(top=max(custom_ticks))
# ax.set_title('Current rose 3D')
# ax.legend()

# ax = WindroseAxes.from_ax()
# ax.bar(current_met_data['heading (u/v)'], current_met_data['Uc'], normed = True, opening = 1, edgecolor = 'black', nsector = 12)
# custom_ticks = [10, 20, 30, 40, 50]
# ax.set_yticks(custom_ticks)
# ax.set_yticklabels([f"{t}%" for t in custom_ticks])
# ax.set_ylim(top=max(custom_ticks))
# ax.set_title('Current Rose tan(speedU/speedV)')
# ax.set_legend()

## lets test the depth calc using the 1/7th power law
current_met_data['Uc depth_53.5'] = (8/7)*(current_met_data['Uc'])*(0.5/53.5)**(1/7)
# current_met_data['Uc depth_53.1'] = (8/7)*(current_met_data['Uc'])*(0.5/53.1)**(1/7)
# current_met_data['Uc depth_55'] = (8/7)*(current_met_data['Uc'])*(0.5/55)**(1/7)
# current_met_data['Uc depth_49'] = (8/7)*(current_met_data['Uc'])*(0.5/49)**(1/7)

# plt.figure() 
# plt.scatter(current_met_data.index, current_met_data['Uc'], color = 'black', label = 'original Uc')
# plt.scatter(current_met_data.index, current_met_data['depth_53.5'], color = 'orange', label = 'current @ 53.5m')
# plt.scatter(current_met_data.index, current_met_data['depth_53.1'], color = 'blue', label = 'current @ 53.1m')
# plt.scatter(current_met_data.index, current_met_data['depth_55'], color = 'green', label = 'current @ 55m')
# plt.scatter(current_met_data.index, current_met_data['depth_49'], color = 'red', label = 'current @ 49m')
# plt.xlabel('data index')
# plt.ylabel('velocity (m/s)')
# plt.title('Comparison of the different depths on the current velocity')
# plt.legend()
# plt.show()

# Waves are measured at 'coming from' but ABSTAB uses 'going to' so it needs to be converted
#initial wave rose
ax = WindroseAxes.from_ax()
ax.bar(wave_met_data['MWD'], wave_met_data['Hm0'], normed = True, opening = 1, edgecolor = 'black')
ax.set_title('initial wave headings (coming from)')
ax.set_legend()

# to convert (coming from to going to) we need to add 180 and then check if its greater than greater than 360 and then subtract 360 if it is
wave_met_data['heading'] = wave_met_data['MWD'] + 180
for i in range(len(wave_met_data)):
    current_head = wave_met_data.loc[i, 'heading']
    if current_head > 360: 
        current_head = current_head - 360
        wave_met_data.loc[i, 'heading'] = current_head
    else: 
        continue
    
ax = WindroseAxes.from_ax()
ax.bar(wave_met_data['heading'], wave_met_data['Hm0'], normed = True, opening = 1, edgecolor = 'black')
ax.set_title('output wave headings (going to)')
ax.set_legend()

total_df = pd.DataFrame()
total_df['Zone'] = 'P2'
total_df['Event'] = float('nan')
total_df['Time'] = float('nan')
total_df['current_rp'] = float('nan')
total_df['current_heading'] = current_met_data['heading (u/v)']
total_df['Uc_start'] = float('nan')
total_df['Uc'] = current_met_data['Uc depth_53.5']
total_df['z_ref'] = 1
total_df['water_level'] = current_met_data['Surface Elevation'] # fluctuation from average water level? 
total_df['wave_rp'] = float('nan')
total_df['wave_heading'] = wave_met_data['MWD']
total_df['Hs_start'] = float('nan')
total_df['Hs max'] = float('nan')
total_df['Hs'] = wave_met_data['Hm0']
total_df['Tp_start'] = float('nan')
total_df['Tp_max'] = float('nan')
total_df['Tp'] = wave_met_data['Tp']
total_df['peak gamma man'] = float('nan')
total_df['Umax_man'] = float('nan')
total_df['Tmax_man'] = float('nan')

total_df.to_csv(export_path, index = False)

# wb = load_workbook(ABSTAB_temp_sheet)
# ws = wb['Metocean']
# for r_idx, row in enumerate(dataframe_to_rows(total_df, index = False, header = False)):
#     for c_idx, value in enumerate(row):
#         ws.cell(row = 2 + r_idx, column = 1 + c_idx, value = value)
        
# wb.save(ABSTAB_export_path)
# print('metocean data is exported')

